"""
Tests for the export handler (core/export_handler.py).
"""
import io
import csv
import zipfile
import pytest

from core.export_handler import ExportHandler


class TestExcelExport:
    """Tests for Excel export."""

    def test_export_returns_bytes_and_filename(self, sample_test_suite):
        handler = ExportHandler()
        content, filename = handler.export_to_excel(sample_test_suite)
        assert isinstance(content, bytes)
        assert filename.endswith(".xlsx")
        assert len(content) > 0

    def test_export_custom_filename(self, sample_test_suite):
        handler = ExportHandler()
        _, filename = handler.export_to_excel(sample_test_suite, filename="custom.xlsx")
        assert filename == "custom.xlsx"

    def test_export_contains_all_tests(self, sample_test_suite):
        handler = ExportHandler()
        content, _ = handler.export_to_excel(sample_test_suite)

        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active

        # Header row + 3 test rows
        assert ws.max_row == 4  # 1 header + 3 tests
        # Check header
        assert ws.cell(1, 1).value == "Test ID"
        # Check first test
        assert ws.cell(2, 1).value == "TC_001"


class TestCsvExport:
    """Tests for CSV export."""

    def test_export_returns_string_and_filename(self, sample_test_suite):
        handler = ExportHandler()
        content, filename = handler.export_to_csv(sample_test_suite)
        assert isinstance(content, str)
        assert filename.endswith(".csv")

    def test_csv_has_header_and_data(self, sample_test_suite):
        handler = ExportHandler()
        content, _ = handler.export_to_csv(sample_test_suite)

        reader = csv.reader(io.StringIO(content))
        rows = list(reader)

        assert rows[0][0] == "Test ID"  # Header
        assert len(rows) == 4  # 1 header + 3 tests
        assert rows[1][0] == "TC_001"


class TestMarkdownExport:
    """Tests for Markdown export."""

    def test_export_returns_string_and_filename(self, sample_test_suite):
        handler = ExportHandler()
        content, filename = handler.export_to_markdown(sample_test_suite)
        assert isinstance(content, str)
        assert filename.endswith(".md")

    def test_markdown_contains_test_info(self, sample_test_suite):
        handler = ExportHandler()
        content, _ = handler.export_to_markdown(sample_test_suite)

        assert "# Test Suite: Login Feature Tests" in content
        assert "TC_001" in content
        assert "TC_002" in content
        assert "**Priority:**" in content
        assert "| Step |" in content

    def test_markdown_summary_table(self, sample_test_suite):
        handler = ExportHandler()
        content, _ = handler.export_to_markdown(sample_test_suite)

        assert "## Summary" in content
        assert "| Priority | Count |" in content


class TestGherkinExport:
    """Tests for Gherkin file export."""

    def test_export_gherkin_files(self, sample_test_suite):
        handler = ExportHandler()
        files = handler.export_gherkin_files(sample_test_suite)
        assert len(files) == 1
        content, filename = files[0]
        assert filename.endswith(".feature")
        assert "Feature:" in content

    def test_adds_extension_if_missing(self):
        from models.test_case import AutomationScript, TestSuite
        suite = TestSuite(
            name="Test",
            gherkin_scripts=[AutomationScript(
                script_type="gherkin", filename="login", content="Feature: Login"
            )]
        )
        handler = ExportHandler()
        files = handler.export_gherkin_files(suite)
        assert files[0][1] == "login.feature"


class TestSeleniumExport:
    """Tests for Selenium file export."""

    def test_export_selenium_files(self, sample_test_suite):
        handler = ExportHandler()
        files = handler.export_selenium_files(sample_test_suite)
        assert len(files) == 1
        content, filename = files[0]
        assert filename.endswith(".py")
        assert "selenium" in content.lower() or "import" in content


class TestPlaywrightExport:
    """Tests for Playwright file export."""

    def test_export_playwright_files(self, sample_test_suite):
        handler = ExportHandler()
        files = handler.export_playwright_files(sample_test_suite)
        assert len(files) == 1
        content, filename = files[0]
        assert filename.endswith(".spec.js")
        assert "playwright" in content.lower() or "test" in content


class TestZipExport:
    """Tests for ZIP bundle export."""

    def test_zip_contains_all_artifacts(self, sample_test_suite):
        handler = ExportHandler()
        content, filename = handler.export_all_as_zip(sample_test_suite)

        assert filename.endswith(".zip")
        assert isinstance(content, bytes)

        with zipfile.ZipFile(io.BytesIO(content), "r") as zf:
            names = zf.namelist()
            # Should contain manual tests, gherkin, selenium, playwright, README
            assert any("manual_tests/" in n for n in names)
            assert any("gherkin/" in n for n in names)
            assert any("selenium/" in n for n in names)
            assert any("playwright/" in n for n in names)
            assert "README.md" in names

    def test_zip_readme_content(self, sample_test_suite):
        handler = ExportHandler()
        content, _ = handler.export_all_as_zip(sample_test_suite)

        with zipfile.ZipFile(io.BytesIO(content), "r") as zf:
            readme = zf.read("README.md").decode("utf-8")
            assert "Login Feature Tests" in readme
            assert "Manual Tests:" in readme
