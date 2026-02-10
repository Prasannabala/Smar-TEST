"""
Export Handler - Generate files in various formats (Excel, CSV, Markdown).
"""
import io
import csv
import zipfile
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any, Optional, Tuple

from models.test_case import ManualTestCase, AutomationScript, TestSuite
from storage.file_manager import get_file_manager, FileManager


class ExportHandler:
    """
    Handles exporting test cases to various file formats.
    """

    def __init__(self):
        self.file_manager: FileManager = get_file_manager()

    def export_to_excel(self, test_suite: TestSuite, filename: Optional[str] = None) -> Tuple[bytes, str]:
        """
        Export manual test cases to Excel format.

        Args:
            test_suite: TestSuite to export
            filename: Optional custom filename

        Returns:
            Tuple of (file bytes, filename)
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Manual Test Cases"

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell_alignment = Alignment(vertical="top", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Priority colors
        priority_fills = {
            "High": PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid"),
            "Medium": PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid"),
            "Low": PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"),
        }

        # Headers (Standard QA format)
        headers = [
            "Test ID", "Test Name", "Description", "Preconditions",
            "Test Steps", "Expected Results", "Priority", "Status", "Category", "Tags"
        ]

        # Column widths
        column_widths = [10, 35, 40, 30, 50, 40, 10, 12, 15, 20]

        # Write headers
        for col, (header, width) in enumerate(zip(headers, column_widths), 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col)].width = width

        # Write test cases
        for row, test in enumerate(test_suite.manual_tests, 2):
            # Test ID
            ws.cell(row=row, column=1, value=test.test_id).border = thin_border

            # Test Name
            ws.cell(row=row, column=2, value=test.test_name).border = thin_border

            # Description
            ws.cell(row=row, column=3, value=test.description).border = thin_border

            # Preconditions
            preconditions = test.get_preconditions_text()
            ws.cell(row=row, column=4, value=preconditions).border = thin_border

            # Test Steps
            steps = test.get_steps_text()
            ws.cell(row=row, column=5, value=steps).border = thin_border

            # Expected Results
            expected = test.get_expected_results_text()
            ws.cell(row=row, column=6, value=expected).border = thin_border

            # Priority (with color)
            priority_cell = ws.cell(row=row, column=7, value=test.priority)
            priority_cell.border = thin_border
            if test.priority in priority_fills:
                priority_cell.fill = priority_fills[test.priority]

            # Status
            ws.cell(row=row, column=8, value=test.status).border = thin_border

            # Category
            ws.cell(row=row, column=9, value=test.category).border = thin_border

            # Tags
            tags = ", ".join(test.tags) if test.tags else ""
            ws.cell(row=row, column=10, value=tags).border = thin_border

            # Apply alignment to all cells in row
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).alignment = cell_alignment

        # Freeze header row
        ws.freeze_panes = "A2"

        # Auto-filter
        ws.auto_filter.ref = ws.dimensions

        # Generate filename
        if not filename:
            filename = self.file_manager.generate_export_filename(
                test_suite.client_name or "NoClient",
                test_suite.requirement_source or "Tests",
                "ManualTests",
                "xlsx"
            )

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return output.getvalue(), filename

    def export_to_csv(self, test_suite: TestSuite, filename: Optional[str] = None) -> Tuple[str, str]:
        """
        Export manual test cases to CSV format.

        Args:
            test_suite: TestSuite to export
            filename: Optional custom filename

        Returns:
            Tuple of (csv content string, filename)
        """
        output = io.StringIO()
        writer = csv.writer(output)

        # Headers
        headers = [
            "Test ID", "Test Name", "Description", "Preconditions",
            "Test Steps", "Expected Results", "Priority", "Status", "Category", "Tags"
        ]
        writer.writerow(headers)

        # Data rows
        for test in test_suite.manual_tests:
            row = [
                test.test_id,
                test.test_name,
                test.description,
                test.get_preconditions_text(),
                test.get_steps_text(),
                test.get_expected_results_text(),
                test.priority,
                test.status,
                test.category,
                ", ".join(test.tags) if test.tags else ""
            ]
            writer.writerow(row)

        # Generate filename
        if not filename:
            filename = self.file_manager.generate_export_filename(
                test_suite.client_name or "NoClient",
                test_suite.requirement_source or "Tests",
                "ManualTests",
                "csv"
            )

        return output.getvalue(), filename

    def export_to_markdown(self, test_suite: TestSuite, filename: Optional[str] = None) -> Tuple[str, str]:
        """
        Export manual test cases to Markdown format.

        Args:
            test_suite: TestSuite to export
            filename: Optional custom filename

        Returns:
            Tuple of (markdown content, filename)
        """
        lines = [
            f"# Test Suite: {test_suite.name}",
            "",
            f"**Client:** {test_suite.client_name or 'N/A'}",
            f"**Source:** {test_suite.requirement_source or 'N/A'}",
            f"**Generated:** {test_suite.generated_at or datetime.now().isoformat()}",
            f"**Total Tests:** {len(test_suite.manual_tests)}",
            "",
            "---",
            "",
        ]

        # Summary table
        lines.append("## Summary")
        lines.append("")
        lines.append("| Priority | Count |")
        lines.append("|----------|-------|")

        priority_counts = {"High": 0, "Medium": 0, "Low": 0}
        for test in test_suite.manual_tests:
            if test.priority in priority_counts:
                priority_counts[test.priority] += 1

        for priority, count in priority_counts.items():
            lines.append(f"| {priority} | {count} |")

        lines.append("")
        lines.append("---")
        lines.append("")

        # Test cases
        lines.append("## Test Cases")
        lines.append("")

        for test in test_suite.manual_tests:
            lines.append(f"### {test.test_id}: {test.test_name}")
            lines.append("")
            lines.append(f"**Priority:** {test.priority} | **Category:** {test.category} | **Status:** {test.status}")
            lines.append("")
            lines.append(f"**Description:** {test.description}")
            lines.append("")

            if test.preconditions:
                lines.append("**Preconditions:**")
                for pre in test.preconditions:
                    lines.append(f"- {pre}")
                lines.append("")

            lines.append("**Test Steps:**")
            lines.append("")
            lines.append("| Step | Action | Test Data | Expected Result |")
            lines.append("|------|--------|-----------|-----------------|")
            for step in test.test_steps:
                lines.append(f"| {step.step_number} | {step.action} | {step.test_data or '-'} | {step.expected_result or '-'} |")
            lines.append("")

            if test.expected_results:
                lines.append("**Expected Results:**")
                for i, result in enumerate(test.expected_results, 1):
                    lines.append(f"{i}. {result}")
                lines.append("")

            if test.tags:
                lines.append(f"**Tags:** {', '.join(test.tags)}")
                lines.append("")

            lines.append("---")
            lines.append("")

        # Generate filename
        if not filename:
            filename = self.file_manager.generate_export_filename(
                test_suite.client_name or "NoClient",
                test_suite.requirement_source or "Tests",
                "ManualTests",
                "md"
            )

        return "\n".join(lines), filename

    def export_gherkin_files(self, test_suite: TestSuite) -> List[Tuple[str, str]]:
        """
        Export Gherkin feature files.

        Args:
            test_suite: TestSuite containing Gherkin scripts

        Returns:
            List of (content, filename) tuples
        """
        files = []
        for script in test_suite.gherkin_scripts:
            filename = script.filename
            if not filename.endswith('.feature'):
                filename += '.feature'
            files.append((script.content, filename))
        return files

    def export_selenium_files(self, test_suite: TestSuite) -> List[Tuple[str, str]]:
        """
        Export Selenium Python test files.

        Args:
            test_suite: TestSuite containing Selenium scripts

        Returns:
            List of (content, filename) tuples
        """
        files = []
        for script in test_suite.selenium_scripts:
            filename = script.filename
            if not filename.endswith('.py'):
                filename += '.py'
            files.append((script.content, filename))
        return files

    def export_playwright_files(self, test_suite: TestSuite) -> List[Tuple[str, str]]:
        """
        Export Playwright JavaScript test files.

        Args:
            test_suite: TestSuite containing Playwright scripts

        Returns:
            List of (content, filename) tuples
        """
        files = []
        for script in test_suite.playwright_scripts:
            filename = script.filename
            if not filename.endswith('.spec.js'):
                if filename.endswith('.js'):
                    filename = filename[:-3] + '.spec.js'
                else:
                    filename += '.spec.js'
            files.append((script.content, filename))
        return files

    def export_all_as_zip(self, test_suite: TestSuite, export_format: str = "excel") -> Tuple[bytes, str]:
        """
        Export all test artifacts as a ZIP bundle.

        Args:
            test_suite: TestSuite to export
            export_format: Format for manual tests ("excel", "csv", or "markdown")

        Returns:
            Tuple of (zip bytes, filename)
        """
        output = io.BytesIO()

        with zipfile.ZipFile(output, 'w', zipfile.ZIP_DEFLATED) as zipf:
            # Export manual tests
            if export_format == "excel":
                content, fname = self.export_to_excel(test_suite)
                zipf.writestr(f"manual_tests/{fname}", content)
            elif export_format == "csv":
                content, fname = self.export_to_csv(test_suite)
                zipf.writestr(f"manual_tests/{fname}", content)
            else:  # markdown
                content, fname = self.export_to_markdown(test_suite)
                zipf.writestr(f"manual_tests/{fname}", content)

            # Export Gherkin files
            for content, fname in self.export_gherkin_files(test_suite):
                zipf.writestr(f"gherkin/{fname}", content)

            # Export Selenium files
            for content, fname in self.export_selenium_files(test_suite):
                zipf.writestr(f"selenium/{fname}", content)

            # Export Playwright files
            for content, fname in self.export_playwright_files(test_suite):
                zipf.writestr(f"playwright/{fname}", content)

            # Add a README
            readme = self._generate_readme(test_suite)
            zipf.writestr("README.md", readme)

        output.seek(0)

        # Generate ZIP filename
        zip_filename = self.file_manager.generate_export_filename(
            test_suite.client_name or "NoClient",
            test_suite.requirement_source or "Tests",
            "TestSuite",
            "zip"
        )

        return output.getvalue(), zip_filename

    def _generate_readme(self, test_suite: TestSuite) -> str:
        """Generate README for the ZIP bundle."""
        summary = test_suite.get_summary()

        lines = [
            f"# {test_suite.name}",
            "",
            f"Generated: {test_suite.generated_at}",
            f"Client: {test_suite.client_name or 'N/A'}",
            f"Source: {test_suite.requirement_source}",
            "",
            "## Contents",
            "",
            f"- **Manual Tests:** {summary['manual_tests']} test cases",
            f"- **Gherkin Scenarios:** {summary['gherkin_scenarios']} scenarios",
            f"- **Selenium Tests:** {summary['selenium_tests']} scripts",
            f"- **Playwright Tests:** {summary['playwright_tests']} specs",
            "",
            "## Directory Structure",
            "",
            "```",
            "├── manual_tests/     # Manual test cases",
            "├── gherkin/          # Gherkin feature files",
            "├── selenium/         # Selenium Python tests",
            "├── playwright/       # Playwright JavaScript tests",
            "└── README.md         # This file",
            "```",
            "",
            "## Usage",
            "",
            "### Gherkin",
            "Run with Cucumber or Behave:",
            "```",
            "behave gherkin/",
            "```",
            "",
            "### Selenium",
            "Run with pytest:",
            "```",
            "pytest selenium/",
            "```",
            "",
            "### Playwright",
            "Run with npx:",
            "```",
            "npx playwright test playwright/",
            "```",
        ]

        return "\n".join(lines)


# Factory function
def get_export_handler() -> ExportHandler:
    """Get export handler instance."""
    return ExportHandler()
